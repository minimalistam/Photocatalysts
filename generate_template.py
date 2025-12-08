import pandas as pd
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils.dataframe import dataframe_to_rows
from pathlib import Path

def generate_excel_template():
    """Generates an Excel template with dropdown validation."""
    
    # 1. Define Columns and Example Data
    columns = [
        "A_element", "A_oxidation", 
        "B_element", "B_oxidation", 
        "X_element", "X_oxidation", 
        "synthesis_temperature", "synthesis_time_hours", 
        "annealing_temperature", "annealing_time_hours", 
        "crystal_structure", "sample_form", 
        "synthesis_method", "morphology", 
        "bandgap_type", "bandgap_method",
        "lattice_parameter", "crystallite_size" # Added for Spinel support
    ]
    
    # Create an empty DataFrame with these columns
    df = pd.DataFrame(columns=columns)
    
    # Add a few example rows (optional, but helpful)
    examples = [
        {
            "A_element": "Cs", "A_oxidation": 1, "B_element": "Pb", "B_oxidation": 2, "X_element": "I", "X_oxidation": -1,
            "synthesis_temperature": 100, "synthesis_time_hours": 2, "annealing_temperature": 0, "annealing_time_hours": 0,
            "crystal_structure": "Cubic", "sample_form": "Powder", "synthesis_method": "Solution", "morphology": "Unknown",
            "bandgap_type": "Direct"
        },
        {
            "A_element": "Mg", "A_oxidation": 2, "B_element": "Al", "B_oxidation": 3, "X_element": "O", "X_oxidation": -2,
            "synthesis_temperature": 600, "synthesis_time_hours": 12, "lattice_parameter": 8.08, "crystallite_size": 20,
            "crystal_structure": "Cubic", "sample_form": "Powder", "synthesis_method": "Solid-State", "morphology": "Bulk",
            "bandgap_type": "Direct"
        }
    ]
    
    df = pd.concat([df, pd.DataFrame(examples)], ignore_index=True)

    # 2. Create Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Batch Prediction Template"

    # Write DataFrame to Worksheet
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    # 3. Define Validations
    validations = {
        "crystal_structure": ["Cubic", "Tetragonal", "Orthorhombic"],
        "sample_form": ["Powder", "Film", "Single Crystal"],
        "synthesis_method": ["Solution", "Solid-State", "Combustion", "Hydrothermal", "Sol-Gel", "Co-precipitation", "Unknown"],
        "morphology": ["Nanoparticles", "Bulk", "Film", "Spherical", "Agglomerated", "Unknown"],
        "bandgap_type": ["Direct", "Indirect"],
        "bandgap_method": ["Tauc plot", "UV-Vis", "DRS", "Unknown"]
    }

    # Helper to find column letter
    def get_col_letter(col_name):
        try:
            idx = columns.index(col_name) + 1
            from openpyxl.utils import get_column_letter
            return get_column_letter(idx)
        except ValueError:
            return None

    # Apply Validations
    for col_name, options in validations.items():
        col_letter = get_col_letter(col_name)
        if col_letter:
            # Create the validation object
            # Formula needs to be a comma-separated string in quotes
            formula = f'"{",".join(options)}"'
            dv = DataValidation(type="list", formula1=formula, allow_blank=True)
            
            # Add validation to the worksheet
            ws.add_data_validation(dv)
            
            # Apply to rows 2 to 1000
            dv.add(f"{col_letter}2:{col_letter}1000")

    # 4. Save
    output_path = Path("data/batch_template.xlsx")
    output_path.parent.mkdir(exist_ok=True)
    wb.save(output_path)
    print(f"Created template at {output_path}")

if __name__ == "__main__":
    generate_excel_template()
